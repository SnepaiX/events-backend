from django.http import HttpResponse
from django.shortcuts import render
from openpyxl.workbook import Workbook
from rest_framework.viewsets import ModelViewSet
from rest_framework.permissions import IsAdminUser, AllowAny
from django.utils import timezone

from .models import Event, EventImage
from .serializers import EventSerializer, EventImageSerializer
from .filters import EventFilter
from .serializers import ImportXlsxSerializer
from places import models
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.parsers import MultiPartParser, FormParser
from drf_spectacular.utils import extend_schema, OpenApiTypes
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from datetime import datetime


class EventViewSet(ModelViewSet):

    queryset = Event.objects.all()
    serializer_class = EventSerializer

    filter_backends = [
        DjangoFilterBackend,
        SearchFilter,
        OrderingFilter,
    ]

    filterset_class = EventFilter

    search_fields = [
        'name',
        'place__name',
    ]

    ordering_fields = [
        'name',
        'start_at',
        'end_at',
    ]

    # parser_classes = [
    #     MultiPartParser,
    #     FormParser,
    # ]

    ordering = ['start_at']

    def get_permissions(self):

        if self.action in ['list', 'retrieve']:
            return [AllowAny()]

        return [IsAdminUser()]

    def get_queryset(self):

        qs = super().get_queryset()

        user = self.request.user

        if not user.is_staff:
            qs = qs.filter(
                status=Event.STATUS_PUBLISHED,
                publish_at__lte=timezone.now()
            )

        return qs

    def perform_create(self, serializer):

        serializer.save(author=self.request.user)

    @extend_schema(
        request=ImportXlsxSerializer,
        responses={200: dict}
    )
    @action(
        detail=False,
        methods=['post'],
        parser_classes=[MultiPartParser, FormParser],
        url_path='import-xlsx'
    )
    def import_xlsx(self, request):

        file = request.FILES.get("file")

        if not file:
            return Response(
                {"error": "No file provided"},
                status=status.HTTP_400_BAD_REQUEST
            )

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        created = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            (
                name,
                description,
                publish_at,
                start_at,
                end_at,
                place_name,
                latitude,
                longitude,
                rating,
            ) = row

            place, _ = models.Place.objects.get_or_create(
                name=place_name,
                defaults={
                    "latitude": latitude,
                    "longitude": longitude,
                }
            )

            Event.objects.create(
                name=name,
                description=description,
                publish_at=publish_at,
                start_at=start_at,
                end_at=end_at,
                place=place,
                rating=rating,
                author=request.user,
                status=Event.STATUS_DRAFT,
            )

            created += 1

        return Response(
            {"created": created},
            status=status.HTTP_201_CREATED
        )

    @action(
        detail=False,
        methods=["get"],
        url_path="export-xlsx"
    )
    def export_xlsx(self, request):

        wb = Workbook()
        ws = wb.active
        ws.title = "Events"

        # Заголовки
        ws.append([
            "ID",
            "Name",
            "Description",
            "Start",
            "End",
            "Status",
            "Rating",
            "Place",
            "Author",
            "Created",
        ])

        queryset = self.get_queryset()

        for event in queryset:
            ws.append([
                event.id,
                event.name,
                event.description,
                event.start_at.strftime("%Y-%m-%d %H:%M"),
                event.end_at.strftime("%Y-%m-%d %H:%M"),
                event.status,
                event.rating,
                event.place.name if event.place else "",
                event.author.username if event.author else "",
                event.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response["Content-Disposition"] = (
            'attachment; filename="events.xlsx"'
        )

        wb.save(response)

        return response


@extend_schema(
    request=EventImageSerializer,
    responses=EventImageSerializer
)
class EventImageViewSet(ModelViewSet):
    queryset = EventImage.objects.all()
    serializer_class = EventImageSerializer

    parser_classes = [MultiPartParser, FormParser]

